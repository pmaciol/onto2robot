from pathlib import Path
from pprint import pprint

import openpyxl
from owlready2 import Ontology

from onto2robot.core import (
    load_ontology,
)


def import_worksheets_from_excel(filename="KB.xlsx") -> dict[str, list[tuple]]:
    tests_folder = Path(__file__).parent.parent / "tests"
    file_path = tests_folder / filename

    workbook = openpyxl.load_workbook(file_path)

    worksheets = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)

        worksheets[sheet_name] = data

    workbook.close()

    return worksheets


def get_all_premises(ontology: Ontology, imports: dict[str, list[tuple]], drop: int = 2):
    items = imports.get("premises", [])[drop:]
    pprint(items)
    for it in items:
        new_item = ontology.Premise()
        new_item.name = it[0]
        new_item.label = it[0]
        pprint(it[1])
        pprint(ontology[it[1]])

        new_item.hasLeftHand = [ontology[it[1]]]
        new_item.hasRightHand = [ontology[it[2]]]
    return items


def get_all_conclusions(ontology: Ontology, imports: dict[str, list[tuple]], drop: int = 2):
    items = imports.get("conclusions", [])[drop:]
    for it in items:
        new_item = ontology.Conclusion()
        new_item.name = it[0]
        new_item.label = it[0]
        new_item.hasLeftHand = [ontology[it[1]]]
        new_item.hasRightHand = [ontology[it[2]]]
    return items


def get_rulesests(ontology: Ontology, imports: dict[str, list[tuple]], drop: int = 2):
    items = imports.get("rulesSets", [])[drop:]
    r_id = set([it[0] for it in items])
    rulesets = {}
    for rid in r_id:
        contains = [it[1] for it in items if it[0] == rid]
        rulesets[rid] = {"contains": contains}

    for k, v in rulesets.items():
        new_item = ontology.RulesSets()
        new_item.name = k
        new_item.label = k
        print(f"Creating RulesSet {k} with contents {v} {new_item}")
        for p in v["contains"]:
            item = ontology[p]
            print(f" - adding rule {p} -> {item}")
            new_item.contains.append(item)
            print(f" -- now contains {new_item.contains}")
    return items


def get_rules(ontology: Ontology, imports: dict[str, list[tuple]], drop: int = 2):
    items = imports.get("rules", [])[drop:]
    r_id = set([it[0] for it in items])
    rules = {}
    for rid in r_id:
        conclusion = [it[1] for it in items if it[0] == rid]
        premises = [it[2] for it in items if it[0] == rid]
        rules[rid] = {"conclusion": conclusion, "premises": premises}

    for k, v in rules.items():
        new_item = ontology.RuleHeader()
        new_item.name = k
        new_item.label = k
        new_item.hasConclusion = [ontology[v["conclusion"][0]]]
        new_item.hasPremise = [ontology[p] for p in v["premises"]]
    return items


def load():
    return load_ontology("mobile_robot_ontology")


def run_import():
    ontology = load_ontology("mobile_robot_ontology")
    imports = import_worksheets_from_excel()
    rulesets = get_rulesests(ontology, imports)
    pprint(rulesets)
    ontology.save(file="mobile_robot_ontology_import_test.owl", format="rdfxml")
